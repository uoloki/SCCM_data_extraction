import pyodbc
import pandas as pd
from openpyxl.utils import get_column_letter
import logging

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def read_credentials(file_path):
    """Read database credentials from a file."""
    try:
        credentials = {}
        with open(file_path, 'r') as file:
            for line in file:
                key, value = line.strip().split('=')
                credentials[key.strip()] = value.strip()
        logging.info("Successfully read credentials from file")
        return credentials
    except Exception as e:
        logging.error(f"Error reading credentials from file: {e}")
        raise

def create_connection_string(credentials):
    """Create a connection string from credentials."""
    try:
        connection_string = (
            f'DRIVER={credentials["driver"]};'
            f'SERVER={credentials["server"]};'
            f'DATABASE={credentials["database"]};'
            f'UID={credentials["username"]};'
            f'PWD={credentials["password"]}'
        )
        logging.info("Successfully created connection string")
        return connection_string
    except KeyError as e:
        logging.error(f"Missing key in credentials: {e}")
        raise

def execute_query(connection, query):
    """Execute a SQL query and return the results as a DataFrame."""
    try:
        cursor = connection.cursor()
        cursor.execute(query)
        rows = cursor.fetchall()
        columns = [column[0] for column in cursor.description]
        data = pd.DataFrame.from_records(rows, columns=columns)
        cursor.close()
        logging.info("Successfully executed query and fetched data")
        return data
    except Exception as e:
        logging.error(f"Error executing query: {e}")
        raise

def save_to_excel(dataframes, sheetnames, filename):
    """Save DataFrames to an Excel file with adjusted column widths."""
    try:
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            for dataframe, sheetname in zip(dataframes, sheetnames):
                for column in dataframe.columns:
                    dataframe[column + '_Y'] = 'Y'

                dataframe.to_excel(writer, sheet_name=sheetname, index=False)
                worksheet = writer.sheets[sheetname]
                for column_cells in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column_cells[0].column)
                    for cell in column_cells:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except Exception as e:
                            logging.error(f"Error processing cell value: {e}")
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        logging.info(f"Data successfully saved to {filename}")
    except Exception as e:
        logging.error(f"Error saving data to Excel: {e}")
        raise

def main():
    try:
        credentials = read_credentials('credentials.txt')
        connection_string = create_connection_string(credentials)

        queries = {
            'Hardware Inventory': """
                SELECT
                    v_GS_COMPUTER_SYSTEM.Name0 AS ComputerName,
                    v_GS_PROCESSOR.Name0 AS ProcessorName,
                    v_GS_PROCESSOR.NumberOfCores0 AS NumberOfCores,
                    v_GS_X86_PC_MEMORY.TotalPhysicalMemory0 AS TotalPhysicalMemory
                FROM
                    v_GS_COMPUTER_SYSTEM
                JOIN
                    v_GS_PROCESSOR ON v_GS_COMPUTER_SYSTEM.ResourceID = v_GS_PROCESSOR.ResourceID
                JOIN
                    v_GS_X86_PC_MEMORY ON v_GS_COMPUTER_SYSTEM.ResourceID = v_GS_X86_PC_MEMORY.ResourceID
            """,
            'Software Inventory': """
                SELECT
                    v_GS_ADD_REMOVE_PROGRAMS.DisplayName0 AS SoftwareName,
                    v_GS_ADD_REMOVE_PROGRAMS.Version0 AS Version,
                    v_GS_ADD_REMOVE_PROGRAMS.Publisher0 AS Publisher,
                    v_GS_COMPUTER_SYSTEM.Name0 AS ComputerName
                FROM
                    v_GS_ADD_REMOVE_PROGRAMS
                JOIN
                    v_GS_COMPUTER_SYSTEM ON v_GS_ADD_REMOVE_PROGRAMS.ResourceID = v_GS_COMPUTER_SYSTEM.ResourceID
            """,
            'Backup Status': """
                SELECT
                    v_GS_BACKUPSTATUS.BackupDateTime0 AS BackupDateTime,
                    v_GS_BACKUPSTATUS.BackupStatus0 AS BackupStatus,
                    v_GS_COMPUTER_SYSTEM.Name0 AS ComputerName
                FROM
                    v_GS_BACKUPSTATUS
                JOIN
                    v_GS_COMPUTER_SYSTEM ON v_GS_BACKUPSTATUS.ResourceID = v_GS_COMPUTER_SYSTEM.ResourceID
            """
        }

        connection = pyodbc.connect(connection_string)
        logging.info("Successfully connected to the database")

        dataframes = []
        sheetnames = []

        for sheetname, query in queries.items():
            data = execute_query(connection, query)
            dataframes.append(data)
            sheetnames.append(sheetname)

        save_to_excel(dataframes, sheetnames, 'sccm_data.xlsx')

        connection.close()
        logging.info("Database connection closed")
    except Exception as e:
        logging.error(f"Error in main function: {e}")

if __name__ == '__main__':
    main()
