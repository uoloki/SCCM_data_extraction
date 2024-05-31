import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import logging

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def load_and_filter_excel(input_filename, output_filename):
    """Load data from Excel, filter based on _Y columns, and save to new Excel file."""
    try:
        workbook = load_workbook(input_filename)
    except Exception as e:
        logging.error(f"Error loading workbook: {e}")
        return

    try:
        writer = pd.ExcelWriter(output_filename, engine='openpyxl')

        for sheetname in workbook.sheetnames:
            sheet = workbook[sheetname]

            # Load data into a DataFrame
            data = pd.DataFrame(sheet.values)
            data.columns = data.iloc[0]
            data = data[1:]

            # Identify columns with the _Y suffix
            y_columns = [col for col in data.columns if col.endswith('_Y')]

            # For each _Y column, filter the rows where the value is 'Y'
            for y_col in y_columns:
                original_col = y_col[:-2]  # Remove the '_Y' suffix to get the original column name
                data = data[data[y_col] == 'Y']

                # Drop the _Y column
                data = data.drop(columns=[y_col])

            # Save the filtered data to the new Excel file
            data.to_excel(writer, sheet_name=sheetname, index=False)

            # Adjust column widths
            worksheet = writer.sheets[sheetname]
            for column_cells in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column_cells[0].column)
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except Exception as e:
                        logging.error(f"Error processing cell value: {e}")
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        writer.save()
        logging.info(f"Filtered data saved to {output_filename}")
    except Exception as e:
        logging.error(f"Error processing and saving workbook: {e}")

if __name__ == '__main__':
    input_filename = 'sccm_data.xlsx'
    output_filename = 'filtered_sccm_data.xlsx'
    load_and_filter_excel(input_filename, output_filename)
